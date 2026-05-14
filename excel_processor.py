#!/usr/bin/env python3
"""
Agent 通用版：差旅 Excel 自动处理
无GUI、无弹窗，传入文件路径即可自动运行
适合：上传文件 → 自动运行代码 → 返回可下载文件
"""

from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

# ===================== 样式常量（完全保留你的原始代码）=====================
HEADER_FILL = PatternFill(fill_type="solid", start_color="D0CECE", end_color="D0CECE")
ROW_FILL_RETREAT = PatternFill(fill_type="solid", start_color="93D23E", end_color="93D23E")
ROW_FILL_CHANGE = PatternFill(fill_type="solid", start_color="FCC200", end_color="FCC200")
THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
COMMON_FONT = Font(name="Microsoft YaHei", size=10)
COMMON_ALIGN = Alignment(horizontal="center", vertical="center")

# ===================== 工具函数（完全不变）=====================
def safe_drop(df: pd.DataFrame, columns) -> pd.DataFrame:
    return df.drop(columns=[c for c in columns if c in df.columns], errors="ignore")

def move_column(df: pd.DataFrame, col: str, after_col: str) -> pd.DataFrame:
    if col not in df.columns or after_col not in df.columns:
        return df
    cols = list(df.columns)
    cols.remove(col)
    idx = cols.index(after_col) + 1
    cols.insert(idx, col)
    return df[cols]

def to_display_date(series: pd.Series) -> pd.Series:
    parsed = pd.to_datetime(series, errors="coerce")
    output = parsed.dt.strftime("%Y/%m/%d")
    return output.where(parsed.notna(), series.astype(str))

def remove_last_total_row(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    last_row = df.iloc[-1]
    if last_row.astype(str).str.contains("合计", na=False).any():
        return df.iloc[:-1].copy()
    return df

def reset_serial_column(df: pd.DataFrame) -> pd.DataFrame:
    if "序号" in df.columns:
        df = df.copy()
        df["序号"] = range(1, len(df) + 1)
    return df

def style_excel(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column

        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = COMMON_ALIGN
                cell.font = COMMON_FONT
                cell.border = THIN_BORDER

        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(name="Microsoft YaHei", size=10, bold=True)

        header_map = {ws.cell(row=1, column=i).value: i for i in range(1, max_col + 1)}
        order_type_col = header_map.get("订单类型名称") or header_map.get("订单类型")
        if order_type_col:
            for r in range(2, max_row + 1):
                val = ws.cell(row=r, column=order_type_col).value
                text = "" if val is None else str(val)
                fill = None
                if "退" in text:
                    fill = ROW_FILL_RETREAT
                elif "改" in text:
                    fill = ROW_FILL_CHANGE
                if fill:
                    for c in range(1, max_col + 1):
                        ws.cell(row=r, column=c).fill = fill

        for c in range(1, max_col + 1):
            letter = get_column_letter(c)
            max_len = 0
            for r in range(1, max_row + 1):
                value = ws.cell(row=r, column=c).value
                text = "" if value is None else str(value)
                max_len = max(max_len, len(text))
            ws.column_dimensions[letter].width = min(max(10, max_len + 2), 60)
    wb.save(path)

# ===================== 业务转换逻辑（100% 原样保留）=====================
def transform_ticket(df_ticket: pd.DataFrame, df_all: pd.DataFrame) -> pd.DataFrame:
    df = df_ticket.copy()
    if "原票号" in df.columns:
        df = df.sort_values(by="原票号", kind="stable")

    df = safe_drop(df, ["成本中心", "供应单号", "原订单编号", "原票号", "预订日期", "乘机人部门", "其他(CNY)"])
    df = df.rename(
        columns={
            "订单编号": "订单号",
            "承运人-票号": "票号",
            "起飞日期": "起飞时间",
            "航程名称": "航段",
            "舱位": "物理舱位",
            "销售价": "票面价",
            "机建": "基建",
            "改签手续费": "改签费",
            "退票手续费": "退票费",
        }
    )

    df = move_column(df, "起飞时间", "乘机")
    df = move_column(df, "航段", "航班号")
    df = move_column(df, "退票费", "改签费")

    df["机票服务费"] = 0.0
    key_col = "票号" if "票号" in df.columns else ("承运人-票号" if "承运人-票号" in df.columns else None)
    if key_col and {"订单类型名称", "结算金额"}.issubset(df.columns):
        normal_mask = df["订单类型名称"].astype(str) == "国内机票正常单"
        settle_num = pd.to_numeric(df["结算金额"], errors="coerce")
        drop_idx: list[int] = []
        fee_idx: list[int] = []

        normal_groups = df[normal_mask].groupby(key_col, dropna=False).groups
        for _, idxes in normal_groups.items():
            idx_list = list(idxes)
            if len(idx_list) < 2:
                continue
            five_rows = [i for i in idx_list if settle_num.loc[i] == 5]
            non_five_rows = [i for i in idx_list if settle_num.loc[i] != 5]
            if five_rows and non_five_rows:
                drop_idx.append(five_rows[0])
                fee_idx.append(non_five_rows[0])

        if fee_idx:
            df.loc[fee_idx, "机票服务费"] = 5
        if drop_idx:
            df = df.drop(index=drop_idx)

        change_mask = df["订单类型名称"].astype(str) == "国内机票改签单"
        change_groups = df[change_mask].groupby(key_col, dropna=False).groups
        drop_change_zero: list[int] = []
        settle_num_after = pd.to_numeric(df["结算金额"], errors="coerce")
        for _, idxes in change_groups.items():
            idx_list = list(idxes)
            if len(idx_list) < 2:
                continue
            zeros = [i for i in idx_list if settle_num_after[i] == 0]
            if zeros:
                drop_change_zero.append(zeros[0])
        if drop_change_zero:
            df = df.drop(index=drop_change_zero)

        refund_mask = df["订单类型名称"].astype(str) == "国内机票退票单"
        refund_groups = df[refund_mask].groupby(key_col, dropna=False).groups
        drop_refund_zero: list[int] = []
        for _, idxes in refund_groups.items():
            idx_list = list(idxes)
            if len(idx_list) < 2:
                continue
            zeros = [i for i in idx_list if settle_num_after[i] == 0]
            if zeros:
                drop_refund_zero.append(zeros[0])
        if drop_refund_zero:
            df = df.drop(index=drop_refund_zero)

    df["机票服务费"] = pd.to_numeric(df["机票服务费"], errors="coerce").fillna(0)
    if "结算金额" in df.columns:
        df["结算金额"] = pd.to_numeric(df["结算金额"], errors="coerce").fillna(0)
        df["实收实付"] = df["结算金额"] + df["机票服务费"]
    else:
        df["实收实付"] = df["机票服务费"]

    df = safe_drop(
        df,
        [
            "员工自付金额", "预订人", "出差申请单号", "是否超标", "报销单号",
            "违背事项", "客票状态", "是否跨期退改", "收款科目", "乘车人证件号",
            "可抵扣税额", "不可抵扣金额", "开票服务费", "票款", "对账状态", "航空公司",
        ],
    )
    return reset_serial_column(df.reset_index(drop=True))

def transform_hotel(df_hotel: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    df = df_hotel.copy()
    if "原订单编号" in df.columns:
        df = df.sort_values(by="原订单编号", kind="stable")

    df = safe_drop(df, ["产品名称", "原订单编号", "供应单号", "入住人部门", "所在城市"])
    df = df.rename(columns={"订单编号": "订单号", "间夜数": "住宿天数"})

    if "订单类型" in df.columns:
        idx = list(df.columns).index("订单类型")
        df.insert(idx, "对账方", "景鸿商旅")
    else:
        df["对账方"] = "景鸿商旅"

    if "入住日期" in df.columns:
        df["入住日期"] = to_display_date(df["入住日期"])
    if "离店日期" in df.columns:
        df["离店日期"] = to_display_date(df["离店日期"])

    df["服务费"] = 0
    if "结算金额" in df.columns:
        df["结算金额"] = pd.to_numeric(df["结算金额"], errors="coerce").fillna(0)
        if "服务费" in df.columns:
            cols = list(df.columns)
            cols.remove("服务费")
            settle_idx = cols.index("结算金额")
            cols.insert(settle_idx, "服务费")
            df = df[cols]

    df = safe_drop(
        df,
        [
            "销售价", "酒店协议类型", "出差申请单号", "报销单号", "预订人",
            "业务发生时间", "酒店付款类型", "是否超标", "违背事项", "收款科目",
            "开票服务费", "供应发票类型", "票款",
        ],
    )

    if {"结算金额", "员工自付金额"}.issubset(df.columns):
        emp_paid = pd.to_numeric(df["员工自付金额"], errors="coerce").fillna(0)
        mask_personal = df["结算金额"] == emp_paid
        df_personal = df.loc[mask_personal].copy()
        df_main = df.loc[~mask_personal].copy()
    else:
        df_personal = df.iloc[0:0].copy()
        df_main = df.copy()

    df_main = safe_drop(df_main, ["员工自付金额"])
    df_personal = safe_drop(df_personal, ["员工自付金额"])
    df_main = reset_serial_column(df_main.reset_index(drop=True))
    df_personal = reset_serial_column(df_personal.reset_index(drop=True))
    return df_main, df_personal

# ===================== ✅ 核心：修复云端读取 Excel BUG =====================
def process_uploaded_file(uploaded_file):
    # 读取二进制（修复跨平台报错）
    file_bytes = uploaded_file.read()
    xls = BytesIO(file_bytes)

    # 读取工作表 + 每次读完重置指针（核心修复）
    df_all = pd.read_excel(xls, sheet_name="全部", engine="openpyxl")
    xls.seek(0)
    
    df_hotel = pd.read_excel(xls, sheet_name="酒店", engine="openpyxl")
    xls.seek(0)
    
    df_ticket = pd.read_excel(xls, sheet_name="国内机票", engine="openpyxl")

    # 你的原有业务逻辑
    df_hotel = remove_last_total_row(df_hotel)
    df_ticket = remove_last_total_row(df_ticket)
    out_ticket = transform_ticket(df_ticket, df_all)
    out_hotel, out_hotel_personal = transform_hotel(df_hotel)

    # 输出到 tmp 目录（Streamlit 云专用）
    tmp_dir = Path("/tmp")
    
    ticket_path = tmp_dir / "国内机票_转换后.xlsx"
    hotel_path = tmp_dir / "酒店_转换后.xlsx"

    with pd.ExcelWriter(ticket_path, engine="openpyxl") as writer:
        out_ticket.to_excel(writer, sheet_name="国内机票", index=False)
    style_excel(ticket_path)

    with pd.ExcelWriter(hotel_path, engine="openpyxl") as writer:
        out_hotel.to_excel(writer, sheet_name="酒店", index=False)
        out_hotel_personal.to_excel(writer, sheet_name="酒店-个人支付", index=False)
    style_excel(hotel_path)

    return ticket_path, hotel_path
